# Copyright: (c) OpenSpug Organization. https://github.com/openspug/spug
# Copyright: (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
from django.views.generic import View
from django.db.models import F,Count,Q
from django.http.response import HttpResponseBadRequest
from libs import json_response, JsonParser, Argument
from apps.setting.utils import AppSetting
from apps.app.models import Deploy
from apps.schedule.models import Task
from apps.monitor.models import Detection
from apps.account.models import Role
from libs.ssh import SSH, AuthenticationException
from paramiko.ssh_exception import BadAuthenticationType
from libs import human_datetime, AttrDict
from openpyxl import load_workbook
from apps.config.models import *
from apps.host.models import *
from django.core.cache import cache
import socket
from apps.host.tasks import  *
HOSTKEY = 'host_all'
DBKEY = 'db_all'
DBMultiKEY = 'dbmulti'

class HostView(View):
    def get(self, request,tag):
        host_id = request.GET.get('id')
        if host_id:
            if not request.user.has_host_perm(host_id):
                return json_response(error='无权访问该主机，请联系管理员')
            return json_response(Host.objects.get(pk=host_id))
        # hosts = Host.objects.filter(deleted_by_id__isnull=True)
        # cache.delete(HOSTKEY)
        # cache.delete(DBKEY)
        cluster = ClusterConfig.objects.all()
        wz = WorkZone.objects.all()
        zz = Zone.objects.all()
        svbag = Servicebag.objects.all()
        polist = Portlist.objects.all()
        dvpo = DevicePositon.objects.all()
        cuser = ConnctUser.objects.all()
        rest = ResourceType.objects.all()
        pj = ProjectConfig.objects.all()
        env = Environment.objects.all()
        if tag =="host":
            hosts = Host.objects.filter(resource_type=(ResourceType.objects.get(name='主机')).id).all()
            zones = [x['zone'] for x in hosts.order_by('zone').values('zone').distinct()]
            tp = [x['top_project'] for x in hosts.order_by('top_project').values('top_project').distinct()]
            ostp = [x["ostp"] for x in hosts.order_by('ostp').values('ostp').distinct() ]
            res_t = [x['resource_type'] for x in hosts.order_by('resource_type').values('resource_type').distinct()]
            w_z = [x['work_zone'] for x in hosts.order_by('work_zone').values('work_zone').distinct()]
            provider = [x['provider'] for x in hosts.order_by('provider').values('provider').distinct()]
            perms = [x.id for x in hosts] if request.user.is_supper else request.user.host_perms
            content = cache.get(HOSTKEY,{})
            if content:
                return json_response(content)
            content = {"cs": [ x.to_dict() for x in cluster],"wz": [ x.to_dict()for x in wz],
                                  "zz":[ x.to_dict() for x in zz],"svbag":[ x.to_dict() for x in svbag],
                                  "polist":[ x.to_dict()for x in polist],"dvpo":[ x.to_dict()for x in dvpo],
                                  "cuser":[x.to_dict() for x in cuser],"rset": [ x.to_dict() for x in rest],"pj":[x.to_dict() for x in pj],"envs": [x.to_dict() for x in env],
                                  "tp":tp,"ostp":ostp,'provider':provider,'w_z':w_z,'res_t':res_t,'zones': zones, 'hosts': [x.to_dict() for x in hosts], 'perms': perms}
            cache.set(HOSTKEY,content,50*10000)
            return json_response(content)
        else:
            hosts = Host.objects.filter(resource_type=(ResourceType.objects.get(name='数据库')).id).exclude(ipaddress="").all().annotate(num=Count("dbtag")).order_by('-ipaddress')
            zones = [x['zone'] for x in hosts.order_by('zone').values('zone').distinct()]
            tp = [x['top_project'] for x in hosts.order_by('top_project').values('top_project').distinct()]
            ostp = [x["ostp"] for x in hosts.order_by('ostp').values('ostp').distinct()]
            res_t = [x['resource_type'] for x in hosts.order_by('resource_type').values('resource_type').distinct()]
            w_z = [x['work_zone'] for x in hosts.order_by('work_zone').values('work_zone').distinct()]
            provider = [x['provider'] for x in hosts.order_by('provider').values('provider').distinct()]
            perms = [x.id for x in hosts] if request.user.is_supper else request.user.host_perms
            content = cache.get(DBKEY,{})
            if content:
                return json_response(content)
            content = {"cs": [x.to_dict() for x in cluster], "wz": [x.to_dict() for x in wz],
                                  "zz": [x.to_dict() for x in zz], "svbag": [x.to_dict() for x in svbag],
                                  "polist": [x.to_dict() for x in polist], "dvpo": [x.to_dict() for x in dvpo],
                                  "cuser": [x.to_dict() for x in cuser], "rset": [x.to_dict() for x in rest],
                                  "pj": [x.to_dict() for x in pj], "envs": [x.to_dict() for x in env],
                                  "tp": tp, "ostp": ostp, 'provider': provider, 'w_z': w_z, 'res_t': res_t,
                                  'zones': zones, 'hosts': [x.to_dict() for x in hosts], 'perms': perms}

            cache.set(DBKEY,content,50*10000)
            return json_response(content)


    def post(self, request,tag):
        form, error = JsonParser(
            Argument('id', type=int, required=False),
            Argument('zone', type=list, help='请输入资源类别'),
            # Argument('name', help='请输主机名称'),
            Argument('username', type=int, help='请输入登录用户名'),
            Argument('ipaddress', handler=str.strip, help='请输入主机名或IP'),
            Argument('port', type=int, help='请输入SSH端口'),
            Argument('pkey', required=False),
            # Argument('status', handler=str.strip,required=True,help='上线/下线status状态'),
            Argument('comment', required=False),
            Argument('password',handler=str.strip,  required=False),
            Argument('password_expire',type=int, required=False),

            Argument('ostp',type=str, required=False),
            Argument('resource_type',type=int, required=False),
            Argument('v_ip', required=False),
            Argument('outter_ip', required=False),
            Argument('provider',type=int, required=False),
            Argument('top_project', type=list,required=False),
            Argument('child_project', type=list,required=False),
            Argument('service_pack',type=list ,required=False),
            Argument('work_zone',type=int, required=False),
            Argument('use_for', required=False),
            Argument('developer', required=False),
            Argument('opsper', required=False),
            Argument('env_id', type=int, required=False),
            # Argument('ext_config1', required=False),
            Argument('cluster',type=list, required=False),

            Argument('cpus',type=int, required=False),
            Argument('memory', type=float, required=False),
            Argument('osType',handler=str.strip, required=False),
            Argument('osVerion', handler=str.strip, required=False),
            Argument('coreVerion', handler=str.strip, required=False),

            Argument('sys_disk', handler=str.strip, required=False),
            Argument('data_disk', handler=str.strip, required=False),
            Argument('sys_data', handler=str.strip, required=False),
            Argument('dbrelation', type=int, required=False),
            Argument('dbtag', type=str, required=False),

        ).parse(request.body)
        if error is None:
            ppwd = form.pop('password')
            if form.ostp == "Linux" and (ResourceType.objects.get(pk=form.resource_type)).name == "主机":
               if valid_ssh(form.ipaddress, form.port,(ConnctUser.objects.get(pk=form.username)).name, password=ppwd,
                         pkey=form.pkey) is False:
                return json_response('auth fail')
            # form.pop("created_by")
            if tag == "host":
                cache.delete(HOSTKEY)
            if tag == "db":
                cache.delete(DBKEY)

            if form.id:
                pwd = Host.make_password(ppwd)
                # sd = form.pop("sys_disk")
                # ddd = form.pop("data_disk")
                sd = form.sys_disk
                ddd = form.data_disk
                if ddd is not None:
                    if (ddd).find(";") != -1:
                        dd = (ddd).split(";")
                        del dd[-1]
                        one = []
                        two = []
                        for x in dd:
                            if x.find("数据盘:") != -1:
                                one.append(x.split("数据盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])
                        e = []
                        for k, v in dict(zip(one, two)).items():
                            if form.ostp == "Linux":
                                e.append({"type": "xfs", "name": k, "size": v,"used": 0})
                            if form.ostp == "Windows":
                                e.append({"type": "ntfs", "name": k, "size": v,"used": 0})

                        form["data_disk"] = e
                    else:
                        form["data_disk"] = []
                if sd is not None:
                    if form.ostp == "Linux":
                        sd = (form.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                            form["sys_disk"] = [{"type": "xfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                    if form.ostp == "Windows":
                        sd = (form.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                            form["sys_disk"] = [{"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                else:
                    form["sys_disk"] = []
                form['dbtag'] = form.ipaddress + '_V'
                Host.objects.filter(pk=form.pop('id')).update(**form,password_hash=pwd)
                return json_response(error=error)
            # if (ResourceType.objects.get(pk=form.resource_type)).name == "数据库" :
            # union db
            # if Host.objects.filter(ipaddress=form.ipaddress, port=form.port).exists():
            if Host.objects.filter(ipaddress=form.ipaddress,resource_type=(ResourceType.objects.get(pk=form.resource_type)).id).exists():
                    return json_response(error=f'已存在的ip【{form.ipaddress}】')
            else:
                if (ResourceType.objects.get(pk=form.resource_type)).name == "主机" :
                    pwd = Host.make_password(ppwd)
                    sd = form.sys_disk
                    ddd = form.data_disk
                    if ddd is not None:
                        if (ddd).find(";") != -1:
                            dd = (ddd).split(";")
                            del dd[-1]
                            one = []
                            two = []
                            for x in dd:
                                if x.find("数据盘:") != -1:
                                    one.append(x.split("数据盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])
                            e = []
                            for k, v in dict(zip(one, two)).items():
                                if form.ostp == "Linux":
                                    e.append({"type": "xfs", "name": k, "size": v, "used": 0})
                                if form.ostp == "Windows":
                                    e.append({"type": "ntfs", "name": k, "size": v, "used": 0})

                            form["data_disk"] = e
                        else:
                            form["data_disk"] = []
                    if sd is not None:
                        if form.ostp == "Linux":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "xfs", "name": k, "total_size": v, "mount": "/", "used": 0}]

                        if form.ostp == "Windows":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                    else:
                        form["sys_disk"] = []

                    sys_size = [int(x["total_size"]) for x in form["sys_disk"]][0]
                    data_size=0
                    sysdatasize=""
                    if len(form["data_disk"]) > 0:
                        for x in form["data_disk"]:
                            data_size += x["size"]
                    if data_size !=0:
                       sysdatasize = (str(sys_size)+ "G") +"+" + (str(data_size) + "G")
                    else:
                        sysdatasize = str(sys_size) + "G"
                    form["sys_data"] = sysdatasize

                    host = Host.objects.create(create_by=request.user,  password_hash=pwd,**form)
                    update_hostinfo.delay([form.ipaddress], (ConnctUser.objects.get(pk=form.username)).name)
                else:
                    # db not need sysdata
                    pwd = Host.make_password(ppwd)
                    sd = form.pop("sys_disk")
                    ddd = form.pop("data_disk")
                    form['dbtag'] = form.ipaddress + '_V'
                    if ddd is not None:
                        if (ddd).find(";") != -1:
                            dd = (ddd).split(";")
                            del dd[-1]
                            one = []
                            two = []
                            for x in dd:
                                if x.find("数据盘:") != -1:
                                    one.append(x.split("数据盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])
                            e = []
                            for k, v in dict(zip(one, two)).items():
                                if form.ostp == "Linux":
                                    e.append({"type": "xfs", "name": k, "size": v, "used": 0})
                                if form.ostp == "Windows":
                                    e.append({"type": "ntfs", "name": k, "size": v, "used": 0})

                            form["data_disk"] = e
                        else:
                            form["data_disk"] = []
                    if sd is not None:
                        if form.ostp == "Linux":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "xfs", "name": k, "total_size": v, "mount": "/", "used": 0}]

                        if form.ostp == "Windows":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                    else:
                        form["sys_disk"] = []


                    host = Host.objects.create(create_by=request.user, password_hash=pwd, **form)
                if request.user.role:
                    request.user.role.add_host_perm(host.id)
        return json_response(error=error)

    def patch(self, request,tag):
        form, error = JsonParser(
            Argument('id', type=int, required=False),
            Argument('zone', help='请输入主机类别')
        ).parse(request.body)
        if error is None:
            host = Host.objects.filter(pk=form.id).first()
            if not host:
                return json_response(error='未找到指定主机')

            count = Host.objects.filter(zone=host.zone, deleted_by_id__isnull=True).update(zone=form.zone)
            return json_response(count)
        return json_response(error=error)

    def delete(self, request,tag):
        form, error = JsonParser(
            Argument('id', type=int, help='请指定操作对象')
        ).parse(request.GET)
        if error is None:
            deploy = Deploy.objects.filter(host_ids__regex=fr'[^0-9]{form.id}[^0-9]').annotate(
                app_name=F('app__name'),
                env_name=F('env__name')
            ).first()
            if deploy:
                return json_response(error=f'应用【{deploy.app_name}】在【{deploy.env_name}】的发布配置关联了该主机，请解除关联后再尝试删除该主机')
            task = Task.objects.filter(targets__regex=fr'[^0-9]{form.id}[^0-9]').first()
            if task:
                return json_response(error=f'任务计划中的任务【{task.name}】关联了该主机，请解除关联后再尝试删除该主机')
            detection = Detection.objects.filter(type__in=('3', '4'), addr=form.id).first()
            if detection:
                return json_response(error=f'监控中心的任务【{detection.name}】关联了该主机，请解除关联后再尝试删除该主机')
            role = Role.objects.filter(host_perms__regex=fr'[^0-9]{form.id}[^0-9]').first()
            if role:
                return json_response(error=f'角色【{role.name}】的主机权限关联了该主机，请解除关联后再尝试删除该主机')
            trelease = Host.objects.filter(pk=form.id,zone="待回收").exists()
            if trelease:
                return json_response(error=f'主机已待回收')
            if tag == "host":
                cache.delete(HOSTKEY)
            if tag == "db":
                cache.delete(DBKEY)
            t = Host.objects.filter(pk=form.id).first()
            Host.objects.filter(pk=form.id).update(
                # zone="待回收",
                ipaddress="",
                iprelease=t.ipaddress,
                deleted_at=human_datetime(),
                deleted_by=request.user,
            )
        return json_response(error=error)

class MultiDbView(View):
    def get(self, request,tag):
        host_id = request.GET.get('id')
        if host_id:
            if not request.user.has_host_perm(host_id):
                return json_response(error='无权访问该主机，请联系管理员')
            return json_response(MultiDBUser.objects.get(pk=host_id))
        # hosts = Host.objects.filter(deleted_by_id__isnull=True)
        # cache.delete(HOSTKEY)
        # cache.delete(DBKEY)
        cluster = ClusterConfig.objects.all()
        wz = WorkZone.objects.all()
        zz = Zone.objects.all()
        svbag = Servicebag.objects.all()
        polist = Portlist.objects.all()
        dvpo = DevicePositon.objects.all()
        cuser = ConnctUser.objects.all()
        rest = ResourceType.objects.all()
        pj = ProjectConfig.objects.all()
        env = Environment.objects.all()
        if tag =="host":
            hosts = MultiDBUser.objects.filter(resource_type=(ResourceType.objects.get(name='主机')).id).all()
            zones = [x['zone'] for x in hosts.order_by('zone').values('zone').distinct()]
            tp = [x['top_project'] for x in hosts.order_by('top_project').values('top_project').distinct()]
            ostp = [x["ostp"] for x in hosts.order_by('ostp').values('ostp').distinct() ]
            res_t = [x['resource_type'] for x in hosts.order_by('resource_type').values('resource_type').distinct()]
            w_z = [x['work_zone'] for x in hosts.order_by('work_zone').values('work_zone').distinct()]
            provider = [x['provider'] for x in hosts.order_by('provider').values('provider').distinct()]
            perms = [x.id for x in hosts] if request.user.is_supper else request.user.host_perms
            content = cache.get(DBMultiKEY,{})
            if content:
                return json_response(content)
            content = {"cs": [ x.to_dict() for x in cluster],"wz": [ x.to_dict()for x in wz],
                                  "zz":[ x.to_dict() for x in zz],"svbag":[ x.to_dict() for x in svbag],
                                  "polist":[ x.to_dict()for x in polist],"dvpo":[ x.to_dict()for x in dvpo],
                                  "cuser":[x.to_dict() for x in cuser],"rset": [ x.to_dict() for x in rest],"pj":[x.to_dict() for x in pj],"envs": [x.to_dict() for x in env],
                                  "tp":tp,"ostp":ostp,'provider':provider,'w_z':w_z,'res_t':res_t,'zones': zones, 'hosts': [x.to_dict() for x in hosts], 'perms': perms}
            cache.set(DBMultiKEY,content,50*10000)
            return json_response(content)
        else:
            hosts = MultiDBUser.objects.filter(resource_type=(ResourceType.objects.get(name='数据库')).id).exclude(ipaddress="").all()
            zones = [x['zone'] for x in hosts.order_by('zone').values('zone').distinct()]
            tp = [x['top_project'] for x in hosts.order_by('top_project').values('top_project').distinct()]
            ostp = [x["ostp"] for x in hosts.order_by('ostp').values('ostp').distinct()]
            res_t = [x['resource_type'] for x in hosts.order_by('resource_type').values('resource_type').distinct()]
            w_z = [x['work_zone'] for x in hosts.order_by('work_zone').values('work_zone').distinct()]
            provider = [x['provider'] for x in hosts.order_by('provider').values('provider').distinct()]
            perms = [x.id for x in hosts] if request.user.is_supper else request.user.host_perms
            content = cache.get(DBMultiKEY,{})
            if content:
                return json_response(content)
            content = {"cs": [x.to_dict() for x in cluster], "wz": [x.to_dict() for x in wz],
                                  "zz": [x.to_dict() for x in zz], "svbag": [x.to_dict() for x in svbag],
                                  "polist": [x.to_dict() for x in polist], "dvpo": [x.to_dict() for x in dvpo],
                                  "cuser": [x.to_dict() for x in cuser], "rset": [x.to_dict() for x in rest],
                                  "pj": [x.to_dict() for x in pj], "envs": [x.to_dict() for x in env],
                                  "tp": tp, "ostp": ostp, 'provider': provider, 'w_z': w_z, 'res_t': res_t,
                                  'zones': zones, 'hosts': [x.to_dict() for x in hosts], 'perms': perms}

            cache.set(DBMultiKEY,content,50*10000)
            return json_response(content)


    def post(self, request,tag):
        form, error = JsonParser(
            Argument('id', type=int, required=False),
            Argument('zone', type=list, help='请输入资源类别'),
            # Argument('name', help='请输主机名称'),
            Argument('username', type=int, help='请输入登录用户名'),
            Argument('ipaddress', handler=str.strip, help='请输入主机名或IP'),
            Argument('port', type=int, help='请输入SSH端口'),
            Argument('pkey', required=False),
            # Argument('status', handler=str.strip,required=True,help='上线/下线status状态'),
            Argument('comment', required=False),
            Argument('password',handler=str.strip,  required=False),
            Argument('password_expire',type=int, required=False),

            Argument('ostp',type=str, required=False),
            Argument('resource_type',type=int, required=False),
            Argument('v_ip', required=False),
            Argument('outter_ip', required=False),
            Argument('provider',type=int, required=False),
            Argument('top_project', type=list,required=False),
            Argument('child_project', type=list,required=False),
            Argument('service_pack',type=list ,required=False),
            Argument('work_zone',type=int, required=False),
            Argument('use_for', required=False),
            Argument('developer', required=False),
            Argument('opsper', required=False),
            Argument('env_id', type=int, required=False),
            # Argument('ext_config1', required=False),
            Argument('cluster',type=list, required=False),

            Argument('cpus',type=int, required=False),
            Argument('memory', type=float, required=False),
            Argument('osType',handler=str.strip, required=False),
            Argument('osVerion', handler=str.strip, required=False),
            Argument('coreVerion', handler=str.strip, required=False),

            Argument('sys_disk', handler=str.strip, required=False),
            Argument('data_disk', handler=str.strip, required=False),
            Argument('sys_data', handler=str.strip, required=False),

        ).parse(request.body)
        if error is None:
            ppwd = form.pop('password')
            if form.ostp == "Linux" and (ResourceType.objects.get(pk=form.resource_type)).name == "主机":
               if valid_ssh(form.ipaddress, form.port,(ConnctUser.objects.get(pk=form.username)).name, password=ppwd,
                         pkey=form.pkey) is False:
                return json_response('auth fail')
            # form.pop("created_by")
            # if tag == "host":
            #     cache.delete(DBMultiKEY)
            if tag == "db":
                cache.delete(DBMultiKEY)

            if form.id:
                pwd = MultiDBUser.make_password(ppwd)
                # sd = form.pop("sys_disk")
                # ddd = form.pop("data_disk")
                sd = form.sys_disk
                ddd = form.data_disk
                if ddd is not None:
                    if (ddd).find(";") != -1:
                        dd = (ddd).split(";")
                        del dd[-1]
                        one = []
                        two = []
                        for x in dd:
                            if x.find("数据盘:") != -1:
                                one.append(x.split("数据盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])
                        e = []
                        for k, v in dict(zip(one, two)).items():
                            if form.ostp == "Linux":
                                e.append({"type": "xfs", "name": k, "size": v,"used": 0})
                            if form.ostp == "Windows":
                                e.append({"type": "ntfs", "name": k, "size": v,"used": 0})

                        form["data_disk"] = e
                    else:
                        form["data_disk"] = []
                if sd is not None:
                    if form.ostp == "Linux":
                        sd = (form.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                            form["sys_disk"] = [{"type": "xfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                    if form.ostp == "Windows":
                        sd = (form.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                            form["sys_disk"] = [{"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                else:
                    form["sys_disk"] = []

                MultiDBUser.objects.filter(pk=form.pop('id')).update(**form,password_hash=pwd)
                return json_response(error=error)
            # if (ResourceType.objects.get(pk=form.resource_type)).name == "数据库" :
            # union db
            if MultiDBUser.objects.filter(ipaddress=form.ipaddress, port=form.port,username=(ConnctUser.objects.get(pk=form.username)).id).exists():
                return json_response(error=f'已存在的ip【{form.ipaddress}】')
            else:
                if (ResourceType.objects.get(pk=form.resource_type)).name == "主机" :
                    pwd = MultiDBUser.make_password(ppwd)
                    sd = form.sys_disk
                    ddd = form.data_disk
                    if ddd is not None:
                        if (ddd).find(";") != -1:
                            dd = (ddd).split(";")
                            del dd[-1]
                            one = []
                            two = []
                            for x in dd:
                                if x.find("数据盘:") != -1:
                                    one.append(x.split("数据盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])
                            e = []
                            for k, v in dict(zip(one, two)).items():
                                if form.ostp == "Linux":
                                    e.append({"type": "xfs", "name": k, "size": v, "used": 0})
                                if form.ostp == "Windows":
                                    e.append({"type": "ntfs", "name": k, "size": v, "used": 0})

                            form["data_disk"] = e
                        else:
                            form["data_disk"] = []
                    if sd is not None:
                        if form.ostp == "Linux":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "xfs", "name": k, "total_size": v, "mount": "/", "used": 0}]

                        if form.ostp == "Windows":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                    else:
                        form["sys_disk"] = []

                    sys_size = [int(x["total_size"]) for x in form["sys_disk"]][0]
                    data_size=0
                    sysdatasize=""
                    if len(form["data_disk"]) > 0:
                        for x in form["data_disk"]:
                            data_size += x["size"]
                    if data_size !=0:
                       sysdatasize = (str(sys_size)+ "G") +"+" + (str(data_size) + "G")
                    else:
                        sysdatasize = str(sys_size) + "G"
                    form["sys_data"] = sysdatasize

                    host = MultiDBUser.objects.create(create_by=request.user,  password_hash=pwd,**form)
                    update_hostinfo.delay([form.ipaddress], (ConnctUser.objects.get(pk=form.username)).name)
                else:
                    # db not need sysdata
                    pwd = MultiDBUser.make_password(ppwd)
                    sd = form.pop("sys_disk")
                    ddd = form.pop("data_disk")
                    if ddd is not None:
                        if (ddd).find(";") != -1:
                            dd = (ddd).split(";")
                            del dd[-1]
                            one = []
                            two = []
                            for x in dd:
                                if x.find("数据盘:") != -1:
                                    one.append(x.split("数据盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])
                            e = []
                            for k, v in dict(zip(one, two)).items():
                                if form.ostp == "Linux":
                                    e.append({"type": "xfs", "name": k, "size": v, "used": 0})
                                if form.ostp == "Windows":
                                    e.append({"type": "ntfs", "name": k, "size": v, "used": 0})

                            form["data_disk"] = e
                        else:
                            form["data_disk"] = []
                    if sd is not None:
                        if form.ostp == "Linux":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "xfs", "name": k, "total_size": v, "mount": "/", "used": 0}]

                        if form.ostp == "Windows":
                            sd = (form.sys_disk).split(";")
                            one = []
                            two = []
                            for x in sd:
                                if x.find("系统盘:") != -1:
                                    one.append(x.split("系统盘:")[1])

                                if x.find("容量:") != -1:
                                    two.append((x.split("容量:")[1])[:-1])

                            for k, v in dict(zip(one, two)).items():
                                form["sys_disk"] = [
                                    {"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]
                    else:
                        form["sys_disk"] = []


                    host = MultiDBUser.objects.create(create_by=request.user, password_hash=pwd, **form)
                if request.user.role:
                    request.user.role.add_host_perm(host.id)
        return json_response(error=error)

    def patch(self, request,tag):
        form, error = JsonParser(
            Argument('id', type=int, required=False),
            Argument('zone', help='请输入主机类别')
        ).parse(request.body)
        if error is None:
            host = MultiDBUser.objects.filter(pk=form.id).first()
            if not host:
                return json_response(error='未找到指定主机')

            count = MultiDBUser.objects.filter(zone=host.zone, deleted_by_id__isnull=True).update(zone=form.zone)
            return json_response(count)
        return json_response(error=error)

    def delete(self, request,tag):
        form, error = JsonParser(
            Argument('id', type=int, help='请指定操作对象')
        ).parse(request.GET)
        if error is None:
            deploy = Deploy.objects.filter(host_ids__regex=fr'[^0-9]{form.id}[^0-9]').annotate(
                app_name=F('app__name'),
                env_name=F('env__name')
            ).first()
            if deploy:
                return json_response(error=f'应用【{deploy.app_name}】在【{deploy.env_name}】的发布配置关联了该主机，请解除关联后再尝试删除该主机')
            task = Task.objects.filter(targets__regex=fr'[^0-9]{form.id}[^0-9]').first()
            if task:
                return json_response(error=f'任务计划中的任务【{task.name}】关联了该主机，请解除关联后再尝试删除该主机')
            detection = Detection.objects.filter(type__in=('3', '4'), addr=form.id).first()
            if detection:
                return json_response(error=f'监控中心的任务【{detection.name}】关联了该主机，请解除关联后再尝试删除该主机')
            role = Role.objects.filter(host_perms__regex=fr'[^0-9]{form.id}[^0-9]').first()
            if role:
                return json_response(error=f'角色【{role.name}】的主机权限关联了该主机，请解除关联后再尝试删除该主机')
            trelease = MultiDBUser.objects.filter(pk=form.id,zone="待回收").exists()
            if trelease:
                return json_response(error=f'主机已待回收')
            if tag == "dbm":
                cache.delete(DBMultiKEY)
            t = MultiDBUser.objects.filter(pk=form.id).first()
            MultiDBUser.objects.filter(pk=form.id).update(
                # zone="待回收",
                ipaddress="",
                iprelease=t.ipaddress,
                deleted_at=human_datetime(),
                deleted_by=request.user,
            )
        return json_response(error=error)

def post_import(request):
    cache.delete(HOSTKEY)
    cache.delete(DBKEY)
    password = request.POST.get('password')
    file = request.FILES['file']
    ws = load_workbook(file, read_only=True)['Sheet1']
    summary = {'invalid': [], 'skip': [], 'fail': [], 'network': [], 'repeat': [], 'success': [], 'error': []}
    ioctmp = []
    roottmp = []
    for i, row in enumerate(ws.rows):
        if i == 0:  # 第1行是表头 略过
            continue
        # if not all([row[x].value for x in range(5)]):
        #     summary['invalid'].append(i)
        #     continue
        data = AttrDict(
            top_project=row[0].value,
            child_project=row[1].value,
            cluster=row[2].value,
            hostname=row[3].value,
            v_ip=row[4].value,
            outter_ip=row[5].value,
            ipaddress=row[6].value,
            username=row[7].value,
            port=row[8].value,
            password=row[9].value,
            # password_expire=row[7].value,
            zone=row[10].value,
            osType=row[11].value,
            osVerion=row[12].value,
            coreVerion=row[13].value,
            cpus=row[14].value,
            memory=row[15].value,
            ostp=row[16].value,
            provider=row[17].value,
            resource_type=row[18].value,
            sys_disk=row[19].value,
            data_disk=row[20].value,
            work_zone=row[21].value,
            use_for=row[22].value,
            status=row[23].value,
            supplier=row[24].value,
            developer=row[25].value,
            opsper=row[26].value,
            service_pack=row[27].value,
            # host_bug=row[17].value,
            # ext_config1=row[18].value,
            env_id=row[28].value,
            iprelease=row[29].value,
            comment=row[30].value,
            sys_data=row[31].value,
            shili=row[32].value,
            dbrelation=row[33].value,
            dbtag=row[34].value
        )
        uuu = data.username
        if uuu is None:
            summary['skip'].append(i)
            continue
        # if (data.resource_type).strip() == "主机":
        # if Host.objects.filter(ipaddress=data.ipaddress, port=data.port, username=(ConnctUser.objects.get(name=(data.username).strip())).id).exists():
        # if Host.objects.filter(ipaddress=row[i].value, port=row[i].value, username=ConnctUser.objects.get w  ).exists():
            # Host.objects.filter(ipaddress=data.ipaddress, port=data.port, username=data.username).update(create_by=request.user,**data)

        #union
        if Host.objects.filter(ipaddress=data.ipaddress, port=data.port).exists():
            summary['skip'].append(i)
            continue
        # else:
        #     pass
            # if Host.objects.filter(ipaddress=data.ipaddress, port=data.port, username=(ConnctUser.objects.get(name=(data.username).strip())).id).exists():
            #     summary['skip'].append(i)
            #     continue
        try:
            if data.ostp == "" or data.resource_type == "" or data.top_project == "" or data.ipaddress == "" or data.zone == "" \
                    or data.provider == "":
                summary['fail'].append(i)
                continue

            pwd = data.pop('password')
            if data.ostp == 'Linux' and (data.resource_type).strip() == "主机" and valid_ssh(data.ipaddress, data.port, data.username, pwd ) is False:
                summary['fail'].append(i)
                continue
        except AuthenticationException:
            summary['fail'].append(i)
            continue
        except socket.error:
            summary['network'].append(i)
            continue
        except Exception:
            summary['error'].append(i)
            continue
        # if Host.objects.filter(name=data.name, deleted_by_id__isnull=True).exists():
        #     summary['repeat'].append(i)
        #     continue

        # if data.ostp == 'Windows':
        #     pwd = data.pop("password")
        us = data.username
        if (data.resource_type).strip() == "主机":
            sv = []
            if data.service_pack is not None:
                for x in (data.service_pack).split(";"):
                    sv.append((Servicebag.objects.get(name=x)).id)
                data["service_pack"] = sv
            else:
                data["service_pack"] = sv

            # if data.ostp =="Windows":
            if data.data_disk is not None:
                if (data.data_disk).find(";") != -1:
                    dd = (data.data_disk).split(";")
                    del dd[-1]
                    one = []
                    two = []
                    for x in dd:
                        if x.find("数据盘:") != -1:
                            one.append(x.split("数据盘:")[1])

                        if x.find("容量:") != -1:
                            two.append((x.split("容量:")[1])[:-1])
                    e = []

                    for k, v in dict(zip(one, two)).items():
                        if data.ostp == 'Linux':
                            e.append({"type": "xfs", "name": k, "size": v,"used": 0})
                        if data.ostp == 'Windows':
                            e.append({"type": "ntfs", "name": k, "size": v,"used": 0})

                    data["data_disk"] = e
                else:
                    data["data_disk"] = []
            else:
                data["data_disk"] = []

            if data.sys_disk is not None:
                if (data.sys_disk).find(";") !=-1:
                    if data.ostp == 'Linux':
                        sd = (data.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                               data["sys_disk"] = [{"type":"xfs","name":k,"total_size": v,"mount":"/","used": 0}]

                    if data.ostp == 'Windows':
                        sd = (data.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                            data["sys_disk"] = [{"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]

            sys_size = [int(x["total_size"]) for x in data["sys_disk"]][0]
            data_size = 0
            sysdatasize = ""
            if len(data["data_disk"]) > 0:
                for x in data["data_disk"]:
                    data_size += int(x["size"])
            if data_size != 0:
                sysdatasize = (str(sys_size) + "G") + "+" + (str(data_size) + "G")
            else:
                sysdatasize = str(sys_size) + "G"

            data["sys_data"] = sysdatasize


        if (data.resource_type).strip() == "数据库":
            if data['dbrelation'] !="无" and data['dbrelation'] != "" and data['dbrelation'] is not None:
                data['dbtag'] = data['ipaddress'] + '_V'
            if data['dbrelation'] != "" and data['dbrelation'] is not None:
                if data['dbrelation'] == "主":
                    data['dbrelation'] = 1
                if data['dbrelation'] == "从":
                    data['dbrelation'] = 2
                if data['dbrelation'] == "集群":
                    data['dbrelation'] = 3
                if data['dbrelation'] == "无":
                    data['dbrelation'] = 4
            data["sys_disk"] = []
            data["sys_disk"] = []
            data["data_disk"] = []
            data["service_pack"] = []
        if pwd is not None and pwd != "":
            data["password_hash"] = Host.make_password(pwd)
        else:
           data["password_hash"] = None
        data["env_id"] = (Environment.objects.get(name=data.env_id)).id

        tp =[]
        if data.top_project is not None:
            for x in (data.top_project).split(";"):
                tp.append((ProjectConfig.objects.get(name=x)).id)
            data["top_project"] = tp
        else:
            data["top_project"] = tp

        cp = []
        if data.child_project is not None:
            for x in (data.child_project).split(";"):
                cp.append((ProjectConfig.objects.get(name=x)).id)
            data["child_project"] = cp
        else:
            data["child_project"] = cp

        cst = []
        if data.cluster is not None:
            for x in (data.cluster).split(";"):
                cst.append((ClusterConfig.objects.get(name=x)).id)
            data["cluster"] = cst
        else:
            data["cluster"] = cst


        data["username"] = (ConnctUser.objects.get(name=(data.username).strip())).id
        zz = []
        if data.zone is not None:
            for x in (data.zone).split(";"):
                zz.append((Zone.objects.get(name=x)).id)
        rest = data.resource_type

        data["zone"] = zz
        data["provider"] = (DevicePositon.objects.get(name=data.provider)).id
        data["resource_type"] = (ResourceType.objects.get(name=data.resource_type)).id
        data["work_zone"] = (WorkZone.objects.get(name=(data.work_zone).strip())).id

        data["status"] = 0

        host = Host.objects.create(create_by=request.user, **data)
        if rest.strip() == "主机":
            if us == "root":
                roottmp.append(data.ipaddress)
            elif us == "ioc":
                ioctmp.append(data.ipaddress)
        # tmp.append(data.ipaddress)
        # update_hostinfo.delay(tmp, 'root')
        if request.user.role:
            request.user.role.add_host_perm(host.id)
        summary['success'].append(i)
    if len(roottmp) > 0:
        update_hostinfo.delay(roottmp, 'root')
    if len(ioctmp) > 0:
        update_hostinfo.delay(ioctmp, 'ioc')
    return json_response(summary)

def multidb_import(request):
    cache.delete(DBMultiKEY)
    password = request.POST.get('password')
    file = request.FILES['file']
    ws = load_workbook(file, read_only=True)['Sheet1']
    summary = {'invalid': [], 'skip': [], 'fail': [], 'network': [], 'repeat': [], 'success': [], 'error': []}
    # ioctmp = []
    # roottmp = []
    for i, row in enumerate(ws.rows):
        if i == 0:  # 第1行是表头 略过
            continue
        # if not all([row[x].value for x in range(5)]):
        #     summary['invalid'].append(i)
        #     continue
        data = AttrDict(
            top_project=row[0].value,
            child_project=row[1].value,
            cluster=row[2].value,
            hostname=row[3].value,
            v_ip=row[4].value,
            outter_ip=row[5].value,
            ipaddress=row[6].value,
            username=row[7].value,
            port=row[8].value,
            password=row[9].value,
            # password_expire=row[7].value,
            zone=row[10].value,
            osType=row[11].value,
            osVerion=row[12].value,
            coreVerion=row[13].value,
            cpus=row[14].value,
            memory=row[15].value,
            ostp=row[16].value,
            provider=row[17].value,
            resource_type=row[18].value,
            sys_disk=row[19].value,
            data_disk=row[20].value,
            work_zone=row[21].value,
            use_for=row[22].value,
            status=row[23].value,
            supplier=row[24].value,
            developer=row[25].value,
            opsper=row[26].value,
            service_pack=row[27].value,
            # host_bug=row[17].value,
            # ext_config1=row[18].value,
            env_id=row[28].value,
            iprelease=row[29].value,
            comment=row[30].value,
            shili=row[31].value
        )
        uuu = data.username
        if uuu is None:
            summary['skip'].append(i)
            continue
        if (data.resource_type).strip() == "主机":
            if MultiDBUser.objects.filter(ipaddress=data.ipaddress, port=data.port, username=(ConnctUser.objects.get(name=(data.username).strip())).id).exists():
            # if Host.objects.filter(ipaddress=row[i].value, port=row[i].value, username=ConnctUser.objects.get w  ).exists():
                # Host.objects.filter(ipaddress=data.ipaddress, port=data.port, username=data.username).update(create_by=request.user,**data)
                summary['skip'].append(i)
                continue
        else:
            pass
            # if Host.objects.filter(ipaddress=data.ipaddress, port=data.port, username=(ConnctUser.objects.get(name=(data.username).strip())).id).exists():
            #     summary['skip'].append(i)
            #     continue

        try:
            if data.ostp == "" or data.resource_type == "" or data.top_project == "" or data.ipaddress == "" or data.zone == "" \
                    or data.provider == "":
                summary['fail'].append(i)
                continue

            pwd = data.pop('password')
            if data.ostp == 'Linux' and (data.resource_type).strip() == "主机" and valid_ssh(data.ipaddress, data.port, data.username, pwd ) is False:
                summary['fail'].append(i)
                continue
        except AuthenticationException:
            summary['fail'].append(i)
            continue
        except socket.error:
            summary['network'].append(i)
            continue
        except Exception:
            summary['error'].append(i)
            continue
        # if Host.objects.filter(name=data.name, deleted_by_id__isnull=True).exists():
        #     summary['repeat'].append(i)
        #     continue

        # if data.ostp == 'Windows':
        #     pwd = data.pop("password")
        us = data.username
        if (data.resource_type).strip() == "主机":
            sv = []
            if data.service_pack is not None:
                for x in (data.service_pack).split(";"):
                    sv.append((Servicebag.objects.get(name=x)).id)
                data["service_pack"] = sv
            else:
                data["service_pack"] = sv

            # if data.ostp =="Windows":
            if data.data_disk is not None:
                if (data.data_disk).find(";") != -1:
                    dd = (data.data_disk).split(";")
                    del dd[-1]
                    one = []
                    two = []
                    for x in dd:
                        if x.find("数据盘:") != -1:
                            one.append(x.split("数据盘:")[1])

                        if x.find("容量:") != -1:
                            two.append((x.split("容量:")[1])[:-1])
                    e = []

                    for k, v in dict(zip(one, two)).items():
                        if data.ostp == 'Linux':
                            e.append({"type": "xfs", "name": k, "size": v,"used": 0})
                        if data.ostp == 'Windows':
                            e.append({"type": "ntfs", "name": k, "size": v,"used": 0})

                    data["data_disk"] = e
                else:
                    data["data_disk"] = []
            else:
                data["data_disk"] = []

            if data.sys_disk is not None:
                if (data.sys_disk).find(";") !=-1:
                    if data.ostp == 'Linux':
                        sd = (data.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                               data["sys_disk"] = [{"type":"xfs","name":k,"total_size": v,"mount":"/","used": 0}]

                    if data.ostp == 'Windows':
                        sd = (data.sys_disk).split(";")
                        one = []
                        two = []
                        for x in sd:
                            if x.find("系统盘:") != -1:
                                one.append(x.split("系统盘:")[1])

                            if x.find("容量:") != -1:
                                two.append((x.split("容量:")[1])[:-1])

                        for k, v in dict(zip(one, two)).items():
                            data["sys_disk"] = [{"type": "ntfs", "name": k, "total_size": v, "mount": "/", "used": 0}]

            sys_size = [int(x["total_size"]) for x in data["sys_disk"]][0]
            data_size = 0
            sysdatasize = ""
            if len(data["data_disk"]) > 0:
                for x in data["data_disk"]:
                    data_size += int(x["size"])
            if data_size != 0:
                sysdatasize = (str(sys_size) + "G") + "+" + (str(data_size) + "G")
            else:
                sysdatasize = str(sys_size) + "G"

            data["sys_data"] = sysdatasize


        if (data.resource_type).strip() == "数据库":
            data["sys_disk"] = []
            data["data_disk"] = []
            data["service_pack"] = []
        if pwd is not None and pwd != "":
            data["password_hash"] = Host.make_password(pwd)
        else:
           data["password_hash"] = None
        data["env_id"] = (Environment.objects.get(name=data.env_id)).id

        tp =[]
        if data.top_project is not None:
            for x in (data.top_project).split(";"):
                tp.append((ProjectConfig.objects.get(name=x)).id)
            data["top_project"] = tp
        else:
            data["top_project"] = tp

        cp = []
        if data.child_project is not None:
            for x in (data.child_project).split(";"):
                cp.append((ProjectConfig.objects.get(name=x)).id)
            data["child_project"] = cp
        else:
            data["child_project"] = cp

        cst = []
        if data.cluster is not None:
            for x in (data.cluster).split(";"):
                cst.append((ClusterConfig.objects.get(name=x)).id)
            data["cluster"] = cst
        else:
            data["cluster"] = cst


        data["username"] = (ConnctUser.objects.get(name=(data.username).strip())).id
        zz = []
        if data.zone is not None:
            for x in (data.zone).split(";"):
                zz.append((Zone.objects.get(name=x)).id)
        rest = data.resource_type

        data["zone"] = zz
        data["provider"] = (DevicePositon.objects.get(name=data.provider)).id
        data["resource_type"] = (ResourceType.objects.get(name=data.resource_type)).id
        data["work_zone"] = (WorkZone.objects.get(name=(data.work_zone).strip())).id

        data["status"] = 0

        host = MultiDBUser.objects.create(create_by=request.user, **data)
        # if rest.strip() == "主机":
        #     if us == "root":
        #         roottmp.append(data.ipaddress)
        #     elif us == "ioc":
        #         ioctmp.append(data.ipaddress)
        # tmp.append(data.ipaddress)
        # update_hostinfo.delay(tmp, 'root')
        if request.user.role:
            request.user.role.add_host_perm(host.id)
        summary['success'].append(i)
    # if len(roottmp) > 0:
    #     update_hostinfo.delay(roottmp, 'root')
    # if len(ioctmp) > 0:
    #     update_hostinfo.delay(ioctmp, 'ioc')
    return json_response(summary)


def valid_ssh(hostname, port, username, password=None, pkey=None, with_expect=True):
    try:
        private_key = AppSetting.get('private_key')
        public_key = AppSetting.get('public_key')
    except KeyError:
        private_key, public_key = SSH.generate_key()
        AppSetting.set('private_key', private_key, 'ssh private key')
        AppSetting.set('public_key', public_key, 'ssh public key')
    if password:
        _cli = SSH(hostname, port, username, password=str(password))
        _cli.add_public_key(public_key)
    if pkey:
        private_key = pkey
    try:
        cli = SSH(hostname, port, username, private_key)
        cli.ping()
    except BadAuthenticationType:
        if with_expect:
            raise TypeError('该主机不支持密钥认证，请参考官方文档，错误代码：E01')
        return False
    except AuthenticationException:
        if password and with_expect:
            raise ValueError('密钥认证失败，请参考官方文档，错误代码：E02')
        return False
    return True


def post_parse(request):
    file = request.FILES['file']
    if file:
        data = file.read()
        return json_response(data.decode())
    else:
        return HttpResponseBadRequest()

class ModifyPwd(View):
    def post(self,request):
        form, error = JsonParser(
            Argument('data', type=list, required=False)
        ).parse(request.body)
        if error is None:
            # ioctmp = []
            # roottmp = []
            if len(form.data) == 0 :
                item = Host.objects.exclude(ipaddress="19.104.50.128").values("ipaddress","username").all()
                for x in item:
                    u = ConnctUser.objects.get(pk=x["username"]).name
                    if u:
                        if u=="ioc":
                            update_pwd.delay(x['ipaddress'],'ioc')
                            # ioctmp.append(x["ipaddress"])
                        if u == "root":
                            update_pwd.delay(x['ipaddress'],'root')
                            # roottmp.append(x["ipaddress"])
            if len(form.data) > 0 :
                item = Host.objects.filter(pk__in=form.data).exclude(ipaddress="19.104.50.128").values("ipaddress","username").all()
                for x in item:
                    u = ConnctUser.objects.get(pk=x["username"]).name
                    if u:
                        if u == "ioc":
                            update_pwd.delay(x['ipaddress'], 'ioc')
                            # ioctmp.append(x["ipaddress"])
                        if u == "root":
                            update_pwd.delay(x['ipaddress'], 'root')
                            # roottmp.append(x["ipaddress"])
            cache.delete(HOSTKEY)
            # if len(roottmp) > 0:
            #     update_pwd.delay(roottmp, 'root')
            # if len(ioctmp) > 0:
            #     update_pwd.delay(ioctmp, 'ioc')
            return json_response(error=error)
        return json_response(error=error)


